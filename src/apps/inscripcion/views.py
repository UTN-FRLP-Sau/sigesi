# future

# Librerias Standars
from datetime import date

# Librerias de Terceros
import openpyxl

# Django
from .forms import InscripcionForm, SubirDocumentoForm, SubirCertificadoForm
from .models import Inscripcion, Archivos
from django.shortcuts import render, redirect
import threading
from django.conf import settings
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.contrib.auth.mixins import LoginRequiredMixin
from django.core.mail import EmailMultiAlternatives
from django.http import HttpResponseRedirect, HttpResponse, JsonResponse
from django.shortcuts import render
from django.urls import reverse
from django.template.loader import get_template
from django.views.generic.edit import UpdateView, CreateView, FormView

from django.views.generic import View, TemplateView, ListView

# Django Locales
from .forms import CreatePersonaForm, CreateStudentForm, VerificacionInscripcionForm, SubirDocumentoForm, SubirCertificadoForm, ActualizarInscripcionForm, InscripcionForm
from .models import Persona, Estudiante, Archivos, Curso, ESPECIALIDAD_ESTUDIANTE_CHOICES, Turno, ModalidadCursado, Especialidad

# Funciones generales


def create_email(user_mail, subject, template_name, context, request):
    template = get_template(template_name)
    content = template.render(context=context, request=request)

    message = EmailMultiAlternatives(
        subject=subject,
        body='',
        from_email=settings.EMAIL_HOST_USER,
        to=[
            user_mail
        ],
        cc=[]
    )

    message.attach_alternative(content, 'text/html')
    return message

class InscripcionCerrada(TemplateView):
    template_name = "inscripcion/inscripcion_cerrada.html"

#####################################################################################################
#####################################################################################################
#####################################################################################################


class ConfirmacionInformacion(TemplateView):
    template_name = "admin/home.html"

class CreatePersonaAndEstudent(View):
    template_name = 'inscripcion/create_student.html'
    
    def dispatch(self, request, *args, **kwargs):
        hoy = date.today()  # hoy
        cursos = Curso.objects.filter(inscripcion_inicio__lte=hoy, inscripcion_cierre__gte=hoy)
        if cursos.count()==0:
            return HttpResponseRedirect(reverse('preinscripcion_cerrado'))
        else:
            return super().dispatch(request, *args, **kwargs)

    def get(self, request):
        contexto = {
            'persona_form': CreatePersonaForm(),
            'estudiante_form': CreateStudentForm()
        }
        return render(request, self.template_name, contexto)

    def post(self, request):
        persona_form = CreatePersonaForm(request.POST)
        estudiante_form = CreateStudentForm(request.POST)
        if persona_form.is_valid() and estudiante_form.is_valid():
            # Guarda la instancia de persona sin guardar en la base de datos
            persona = persona_form.save(commit=False)
            persona.save()  # Ahora persona tiene un ID asignado
            # Guarda la instancia de estudiante sin guardar en la base de datos
            estudiante = estudiante_form.save(commit=False)
            estudiante.persona = persona  # Asigna la persona al estudiante
            estudiante.save()  # Guarda el estudiante en la base de datos
            try:
                email = create_email(
                    user_mail=persona.correo,
                    subject='Confirmación de Inscripción',
                    template_name='inscripcion/confirmacion_correo.html',
                    context={
                        'nombre': persona.nombres.title(),
                        'apellido': persona.apellidos.upper(),
                        'id_estudiante': estudiante.pk,
                        'carrera': estudiante.get_especialidad_display(),
                        'turno': estudiante.get_turno_display(),
                        'modalidad': estudiante.get_modalidad_display()
                    },
                    request=request
                )
                thread = threading.Thread(target=email.send)
                thread.start()
                return render(self.request, 'inscripcion/success.html')
            except:
                Estudiante.objects.filter(pk=estudiante.pk).delete()
                Persona.objects.filter(pk=persona.pk).delete()
        contexto = {
            'persona_form': persona_form,
            'estudiante_form': estudiante_form
        }
        return render(request, self.template_name, contexto)


class VerificacionInscripcion(View):
    form_class = VerificacionInscripcionForm
    template_name = 'inscripcion/verificar_dni.html'

    def get(self, request, *args, **kwargs):
        form = self.form_class()
        return render(request, self.template_name, {'form': form})

    def post(self, request, *args, **kwargs):
        form = self.form_class(request.POST)
        if form.is_valid():
            dni = form.cleaned_data['dni']
            id_estudiante = self.kwargs['id_estudiante']
            try:
                estudiante = Estudiante.objects.get(credencial=id_estudiante)
                try:
                    usuario = Persona.objects.get(id=estudiante.persona.id)
                    if usuario.numero_documento == dni:
                        # Verificación exitosa, guardamos en la sesión
                        request.session['dni_verificado'] = True
                        request.session['credencial'] = usuario.pk
                        return HttpResponseRedirect(reverse('paso_2', kwargs={'pk': id_estudiante}))
                    else:
                        form.add_error('dni', 'El DNI no coincide con el usuario')
                except Persona.DoesNotExist:
                    form.add_error(None, 'El usuario no existe')
            except Estudiante.DoesNotExist:
                form.add_error(None, 'El usuario no existe')
        return render(request, self.template_name, {'form': form})


class ReenvioLinkdocumentacion(FormView):
    form_class = VerificacionInscripcionForm
    template_name = 'inscripcion/info.html'
    

    def get(self, request, *args, **kwargs):
        form = self.form_class()
        return render(request, self.template_name, {'form': form})

    def post(self, request, *args, **kwargs):
        form = self.form_class(request.POST)
        if form.is_valid():
            dni = form.cleaned_data['dni']
            try:
                persona = Persona.objects.get(numero_documento=dni)
                try:
                    estudiante = Estudiante.objects.get(persona=persona)
                    # Crear el correo
                    email = create_email(
                        user_mail=persona.correo,
                        subject='Re-Confirmación de Inscripción',
                        template_name='inscripcion/reenvio_confirmacion_correo.html',
                        context={
                            'nombre': persona.nombres.title(),
                            'apellido': persona.apellidos.upper(),
                            'id_estudiante': estudiante.pk,
                            'carrera': estudiante.get_especialidad_display(),
                            'turno': estudiante.get_turno_display(),
                            'modalidad': estudiante.get_modalidad_display()
                        },
                        request=request
                    )
                    # Enviar el correo en un hilo separado
                    thread = threading.Thread(target=email.send)
                    thread.start()

                    # Renderizar la página de éxito
                    return render(self.request, 'inscripcion/reenvio_success.html', context={'persona': persona})

                except Estudiante.DoesNotExist:
                    # Si no existe un estudiante asociado
                    return HttpResponseRedirect(reverse('crear_persona'))

            except Persona.DoesNotExist:
                # Si no existe la persona
                return HttpResponseRedirect(reverse('crear_persona'))

        # Si el formulario no es válido, recargar el formulario con errores
        return render(request, self.template_name, {'form': form})
    

class CursosView(ListView):
    model = Curso
    template_name = 'cursos/lista_cursos.html'
    context_object_name = 'cursos'
    
    def dispatch(self, request, *args, **kwargs):
        # Verificar si el usuario ha pasado por VerificarDniView
        if not request.session.get('dni_verificado'):
            id_inscripcion = self.kwargs['pk']
            return HttpResponseRedirect(reverse('verificar_dni', kwargs={'id_estudiante': id_inscripcion}))
        return super().dispatch(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        #Obtengo al Estudiante por su credencial en la url
        estudiante = Estudiante.objects.get(pk=self.kwargs['pk'])
        hoy = date.today() #hoy
        cursos_disponibles = Curso.objects.filter(inscripcion_inicio__lte=hoy, inscripcion_cierre__gte=hoy)
        inscripciones_disponibles = Inscripcion.objects.filter(curso__inscripcion_inicio__lte=hoy, curso__inscripcion_cierre__gte=hoy)

        # Última fecha de cierre (antes o igual a hoy)
        ultima_fecha_cierre = Curso.objects.filter(inscripcion_cierre__lte=hoy).order_by('-inscripcion_cierre').first()
        # Próxima fecha de apertura (posterior a hoy)
        proxima_fecha_apertura = Curso.objects.filter(inscripcion_inicio__gte=hoy).order_by('inscripcion_inicio').first()
        
        # Cursos en los que el usuario ya está inscrito
        cursos_inscriptos = Inscripcion.objects.filter(estudiante=estudiante).filter(curso__fecha_finalizacion__gte=hoy)
        cursos_inscriptos_idlist = inscripciones_disponibles.filter(estudiante=estudiante).values_list('curso', flat=True)

        # Cursos en los que el usuario puede inscribirse (excluyendo los inscriptos)
        cursos_habilitados = cursos_disponibles.exclude(
            id__in=cursos_inscriptos_idlist)

        # Documentacion
        archivos = Archivos.objects.filter(persona=estudiante.persona)
        if archivos.exists():
            documento = archivos.filter(tipo='Identificacion').last()
            certificado = archivos.filter(tipo='Certificado').last()
            if documento and documento.estado == 2:
                context['documento_form'] = SubirDocumentoForm(prefix='documento')
                context['documento_info'] = False
                context['documentacion'] = False
            elif documento and documento.estado == 1:
                context['documento_form'] = False
                context['documento_info'] = 'si'
                context['documentacion'] = True
            elif documento and documento.estado == 0:
                context['documento_form'] = False
                context['documento_info'] = 'no'
                context['documentacion'] = True
            if certificado and certificado.estado == 2:
                context['certificado_info'] = False
                context['certificado_form'] = SubirCertificadoForm(prefix='certificado')
                context['documentacion'] = False
            elif certificado and certificado.estado == 1:
                context['certificado_form'] = False
                context['certificado_info'] = 'si'
                context['documentacion'] = True
            elif certificado and certificado.estado == 0:
                context['certificado_form'] = False
                context['certificado_info'] = 'no'
                context['documentacion'] = True
        else:
            context['documento_form'] = SubirDocumentoForm(prefix='documento')
            context['certificado_form'] = SubirCertificadoForm(prefix='certificado')
            context['documentacion'] = False

        context['cursos_inscriptos'] = cursos_inscriptos
        context['estudiante'] = estudiante
        context['ESPECIALIDAD_ESTUDIANTE_CHOICES'] = ESPECIALIDAD_ESTUDIANTE_CHOICES
        context['cursos_disponibles'] = cursos_habilitados
        context["ultima_fecha_cierre"] = ultima_fecha_cierre if ultima_fecha_cierre else False,
        context["proxima_fecha_apertura"] = proxima_fecha_apertura if proxima_fecha_apertura else False,
        return context


class ActualizarInscripcionView(UpdateView):
    model = Estudiante
    template_name = 'inscripcion/update.html'
    fields = ['especialidad', 'turno', 'modalidad']

    def dispatch(self, request, *args, **kwargs):
        # Verificar si el usuario ha pasado por VerificarDniView
        if not request.session.get('dni_verificado'):
            id_estudiante = request.session.get('credencial')
            return HttpResponseRedirect(reverse('verificar_dni', kwargs={'id_estudiante': id_estudiante}))
        #else:
            #request.session['dni_verificado']=False
            #id_estudiante = self.kwargs['pk']
        return super().dispatch(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # Agregar dos instancias del formulario subirDocumentacion al contexto
        estudiante = Estudiante.objects.get(pk=self.kwargs['pk'])
        context['estudiante'] = estudiante
        context['inscripcion_form'] = InscripcionForm(prefix='inscripcion')
        archivos = Archivos.objects.filter(persona=estudiante.persona)
        if archivos.exists():
            documento = archivos.filter(tipo='Identificacion').last()
            certificado = archivos.filter(tipo='Certificado').last()
            if documento and documento.estado == 2:
                context['documento_form'] = SubirDocumentoForm(
                    prefix='documento')
                context['documento_info'] = False
            elif documento and documento.estado == 1:
                context['documento_form'] = False
                context['documento_info'] = 'si'
            elif documento and documento.estado == 0:
                context['documento_form'] = False
                context['documento_info'] = 'no'
            if certificado and certificado.estado == 2:
                context['certificado_info'] = False
                context['certificado_form'] = SubirCertificadoForm(
                    prefix='certificado')
            elif certificado and certificado.estado == 1:
                context['certificado_form'] = False
                context['certificado_info'] = 'si'
            elif certificado and certificado.estado == 0:
                context['certificado_form'] = False
                context['certificado_info'] = 'no'
        else:
            context['documento_form'] = SubirDocumentoForm(prefix='documento')
            context['certificado_form'] = SubirCertificadoForm(
                prefix='certificado')
        return context

    def post(self, request, *args, **kwargs):
        form_documento = SubirDocumentoForm(
            request.POST, request.FILES, prefix='documento')
        form_certificado = SubirCertificadoForm(
            request.POST, request.FILES, prefix='certificado')
        estudiante = self.get_object()

        # Actualizar los campos del estudiante desde los datos enviados
        estudiante.especialidad = request.POST.get('especialidad')
        estudiante.turno = request.POST.get('turno')
        estudiante.modalidad = request.POST.get('modalidad')
        estudiante.save()

        archivos = Archivos.objects.filter(persona=estudiante.persona)

        # Guardar documento si es válido y necesario
        if not archivos.filter(tipo='Identificacion', estado__in=[0, 1]).exists() and form_documento.is_valid():
            documento = form_documento.save(commit=False)
            documento.tipo = 'Identificacion'
            documento.persona = estudiante.persona
            documento.save()

        # Guardar certificado si es válido y necesario
        if not archivos.filter(tipo='Certificado', estado__in=[0, 1]).exists() and form_certificado.is_valid():
            certificado = form_certificado.save(commit=False)
            certificado.tipo = 'Certificado'
            certificado.persona = estudiante.persona
            certificado.save()

        # Redirigir a una página de éxito o mensaje de éxito si ambos formularios son válidos
        if form_documento.is_valid() and form_certificado.is_valid():
            # Cambia esto a la URL de éxito deseada
            return render(self.request, 'inscripcion/success2.html', {'estudainte_pk': self.kwargs['pk']})

        # Si alguno de los formularios es inválido, devolver a la misma vista con los errores
        return render(request, self.template_name, {
            'form': self.get_form(),
            'documento_form': form_documento,
            'certificado_form': form_certificado,
        })


class CreateInscripcionView(FormView):
    template_name = 'cursos/lista_cursos.html'  # Archivo del template
    form_class = InscripcionForm

    def form_valid(self, form):
        # Obtener datos adicionales del estudiante y el curso
        estudiante_id = self.request.POST.get('estudiante_id')
        curso_id = self.request.POST.get('curso_id')

        estudiante = Estudiante.objects.get(pk=estudiante_id)
        curso = Curso.objects.get(pk=curso_id)

        # Asignar estudiante y curso al formulario antes de guardar
        inscripcion = form.save(commit=False)
        inscripcion.estudiante = estudiante
        inscripcion.curso = curso
        inscripcion.estado = 'inscripto'  # Estado inicial por defecto
        inscripcion.save()
        
        email = create_email(
            user_mail=estudiante.persona.correo,
            subject='Confirmacion de la Inscripción',
            template_name='correos/inscripcion/create.html',
            context={
                'nombre': estudiante.persona.nombres.title(),
                'apellido': estudiante.persona.apellidos.upper(),
                'carrera': inscripcion.especialidad.nombre.title(),
                'turno': inscripcion.turno.nombre.title(),
                'modalidad': inscripcion.modalidad.nombre.title(),
                'seminario': inscripcion.curso.nombre.title(),
                'id_estudiante': estudiante.pk,
            },
            request=self.request
        )
        thread = threading.Thread(target=email.send)
        thread.start()

        # Redirigir tras éxito
        # Define la URL o vista correspondiente
        return HttpResponseRedirect(reverse('paso_2', kwargs={'pk': estudiante_id}))

    def form_invalid(self, form):
        estudiante_id = self.request.POST.get('estudiante_id')
        # Manejo de errores en el formulario
        return HttpResponseRedirect(reverse('paso_2', kwargs={'pk': estudiante_id}))


class UpdateInscripcionView(FormView):
    template_name = 'cursos/lista_cursos.html'  # Archivo del template
    form_class = InscripcionForm

    def form_valid(self, form):
        # Obtener el id de la inscripcion y el estudiante
        estudiante_id = self.request.POST.get('estudiante_id')
        inscripcion_id = self.request.POST.get('inscripcion_id')

        # Obtener la inscripcion y al estudiante
        inscripcion = Inscripcion.objects.get(pk=inscripcion_id)
        estudiante = Estudiante.objects.get(pk=estudiante_id)

        # Actualizar los datos de la inscripcion segun el formulario
        turno = Turno.objects.get(nombre=form.cleaned_data.get('turno'))
        modalidad = ModalidadCursado.objects.get(nombre=form.cleaned_data.get('modalidad'))
        especialidad = form.cleaned_data.get('especialidad', inscripcion.estado)
        inscripcion.especialidad = especialidad
        inscripcion.turno = turno
        inscripcion.modalidad = modalidad
        inscripcion.save()

        email = create_email(
            user_mail=estudiante.persona.correo,
            subject='Actualizacion de la Inscripción',
            template_name='correos/inscripcion/update.html',
            context={
                'nombre': estudiante.persona.nombres.title(),
                'apellido': estudiante.persona.apellidos.upper(),
                'carrera': especialidad.nombre.title(),
                'turno': turno.nombre.title(),
                'modalidad': modalidad.nombre.title(),
                'seminario': inscripcion.curso.nombre.title(),
            },
            request=self.request
        )
        thread = threading.Thread(target=email.send)
        thread.start()
        # Redirigir tras éxito
        # Define la URL o vista correspondiente
        return HttpResponseRedirect(reverse('paso_2', kwargs={'pk': estudiante_id}))

    def form_invalid(self, form):
        estudiante_id = self.request.POST.get('estudiante_id')
        # Manejo de errores en el formulario
        return HttpResponseRedirect(reverse('paso_2', kwargs={'pk': estudiante_id}))


class SubirDoc(FormView):
    template_name = 'cursos/lista_cursos.html'  # Archivo del template
    
    def post(self, request, *args, **kwargs):
    
        estudiante_id = self.request.POST.get('estudiante_id')
        estudiante = Estudiante.objects.get(pk=int(estudiante_id))

        documento_form = SubirDocumentoForm(request.POST, request.FILES, prefix='documento')
        certificado_form = SubirCertificadoForm(request.POST, request.FILES, prefix='certificado')

        archivos = Archivos.objects.filter(persona=estudiante.persona)

        # Guardar documento si es válido y necesario
        print(estudiante)
        if not archivos.filter(tipo='Identificacion', estado__in=[0, 1]).exists() and documento_form.is_valid():
            documento = documento_form.save(commit=False)
            documento.tipo = 'Identificacion'
            documento.persona = estudiante.persona
            documento.save()

        # Guardar certificado si es válido y necesario
        if not archivos.filter(tipo='Certificado', estado__in=[0, 1]).exists() and certificado_form.is_valid():
            certificado = certificado_form.save(commit=False)
            certificado.tipo = 'Certificado'
            certificado.persona = estudiante.persona
            certificado.save()
            
        # Redirigir a una página de éxito o mensaje de éxito si ambos formularios son válidos
        if documento_form.is_valid() and certificado_form.is_valid():
            messages.error(
                request, 'Documentacion presentada exitosamente')
            return HttpResponseRedirect(reverse('paso_2', kwargs={'pk': estudiante_id}))

        # Si alguno de los formularios es inválido, devolver a la misma vista con los errores
        messages.error(request, "Verifica que el archivo que quieres subir sea un PDF.")
        return HttpResponseRedirect(reverse('paso_2', kwargs={'pk': estudiante_id}))

    def form_invalid(self, form):
        # Manejo de errores en el formulario
        return self.render_to_response(self.get_context_data(form=form))


class EstudianteListView(LoginRequiredMixin, ListView):
    model = Estudiante
    template_name = 'estudiantes/list.html'  # Ruta al template
    # Nombre para acceder a la lista de cursos en el template
    context_object_name = 'estudiantes'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # Todos los cursos para el selector
        hoy = date.today()  # hoy
        cursos = Curso.objects.all()
        context['anios'] = list(set(cursos.values_list('año', flat=True)))
        # Obtener el curso seleccionado, si existe en los parámetros de la URL
        curso_id = self.request.GET.get('curso_id') 
        curso_anio = self.request.GET.get('curso_anio')
        letra = self.request.GET.get('letter')
        if curso_anio:
            context['cursos'] = Curso.objects.filter(año=curso_anio)
        else:
            context['cursos'] = Curso.objects.filter(inscripcion_inicio__lte=hoy, inscripcion_cierre__gte=hoy)
        if curso_id:
            curso = Curso.objects.get(id=curso_id)
            inscripciones = Inscripcion.objects.filter(curso=curso)
            if letra:
                inscripciones = inscripciones.filter(estudiante__persona__apellidos__istartswith=letra)
                context['inscripciones'] = inscripciones
            else:
                context['inscripciones'] = inscripciones[:25]
            # Filtra las inscripciones y estudiantes basándose en el curso
            context['inscripciones_botones'] = True
            context['estudiantes'] = None
            context['curso_version'] = curso
            context['inscripcion_form'] = True
        else:
            # Si no se selecciona un curso, muestra estudiantes no inscritos
            inscritos_ids = Inscripcion.objects.values_list(
                'estudiante_id', flat=True)
            context['estudiantes'] = Estudiante.objects.exclude(
                credencial__in=inscritos_ids)
            context['inscripciones'] = False
        return context


@login_required
def aprobar_curso(request):
    if request.method == 'POST':
        inscripcion_id = request.POST.get('inscripcion_id')
        inscripcion = Inscripcion.objects.get(pk=inscripcion_id)
        if inscripcion.estado != 'aprobado': 
            inscripcion.estado='aprobado'
        else:
            inscripcion.estado='inscripto'
        inscripcion.save()
        return JsonResponse({"success": True, "nuevo_estado": inscripcion.estado})
    return JsonResponse({"success": False}, status=400)


def export_user_asistentes_to_excel(request):
    """Exports UserAsistente data to an Excel file."""

    # Create a new workbook and worksheet
    workbook = openpyxl.Workbook()
    worksheet = workbook.active

    # Set column headers
    headers = [
        'Apellido',                 #1
        'Nombre',                   #2
        'Documento',                #3
        'Correo',                   #4
        'Carrera',                  #5
        'Turno',                    #6
        'Modalidad',                #7
        'Curso',                    #8
        'Escuela',                  #9
        'CUE',                      #10
        'Egreso',                   #11
        'Titulo Secundario',        #12
        'Fecha de Nacimiento',      #13
        'Pais Nacimiento',          #14
        'Pais de Domicilio',        #15
        'Localidad de Nacimiento',  #16
    ]
    for col, header in enumerate(headers, 1):
        worksheet.cell(row=1, column=col, value=header)

    # Fetch UserAsistente data
    user_asistentes = Estudiante.objects.all()

    # Populate worksheet with data
    row = 2
    for asistente in user_asistentes:
        worksheet.cell(row=row, column=1,
                       value=asistente.persona.apellidos.upper())  # Apellido
        worksheet.cell(row=row, column=2,
                       value=asistente.persona.nombres.title())  # Nombre
        worksheet.cell(row=row, column=3,
                       value=asistente.persona.numero_documento)  # Documento
        worksheet.cell(row=row, column=4,
                       value=asistente.persona.correo)  # Correo
        if Inscripcion.objects.filter(estudiante=asistente).order_by('id').last():
            inscripcion = Inscripcion.objects.filter(
                estudiante=asistente).order_by('id').last()
            worksheet.cell(row=row, column=5,
                           value=inscripcion.especialidad.nombre.title())
            worksheet.cell(row=row, column=6,
                           value=inscripcion.turno.nombre.title())
            worksheet.cell(row=row, column=7,
                           value=inscripcion.modalidad.nombre.title())
            worksheet.cell(row=row, column=8, value='{} - ({})'.format(
                inscripcion.curso.nombre.title(), inscripcion.curso.año))
        else:
            worksheet.cell(row=row, column=5,
                           value=asistente.get_especialidad_display())
            worksheet.cell(row=row, column=6,
                           value=asistente.get_turno_display())
            worksheet.cell(row=row, column=7,
                           value=asistente.get_modalidad_display())
            worksheet.cell(row=row, column=8, value='---')
        try:
            worksheet.cell(row=row, column=9, value=asistente.escuela.nombre)
        except:
            worksheet.cell(row=row, column=9, value="---")
        worksheet.cell(row=row, column=10, value=asistente.escuela.cue)
        worksheet.cell(row=row, column=11, value=asistente.anio_egreso)
        worksheet.cell(row=row, column=12, value=asistente.titulo_secundario)
        worksheet.cell(row=row, column=13, value=asistente.persona.fecha_nacimiento)
        worksheet.cell(row=row, column=14, value=asistente.persona.pais_nacimiento.nombre)
        worksheet.cell(row=row, column=15, value=asistente.persona.domicilio_pais.nombre)
        worksheet.cell(row=row, column=16, value=asistente.persona.domicilio_localidad.nombre)
        row += 1

    # Crear una respuesta HTTP con tipo de contenido de Excel
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="inscriptos.xlsx"'
    # Guardar el archivo Excel en la respuesta
    workbook.save(response)

    return response

@login_required
def inscribir_curso(request):
    if request.method == 'POST':
        curso_id = request.POST.get('curso_id')
        documento = request.POST.get('documento')
        aprobacion = request.POST.get('aprobacion')
        curso = Curso.objects.get(pk=curso_id)
        docu_list = str(documento).split(";")
        mensajes =[]
        for docu in docu_list:
            try:
                persona = Persona.objects.get(numero_documento=docu)
                try:
                    estudiante = Estudiante.objects.get(persona=persona)
                    if not Inscripcion.objects.filter(curso=curso).filter(estudiante=estudiante).exists():

                        if str(estudiante.turno) in ['t','m','n']:
                            turno_dic ={'t':2,'m':1,'n':3}
                            turno = turno_dic[str(estudiante.turno)]
                        else:
                            turno = estudiante.turno.pk
                        
                        if str(estudiante.modalidad) in ['p','l','s']:
                            modalidad_dic ={'p':1,'s':2,'l':3}
                            modalidad = modalidad_dic[str(estudiante.modalidad)]
                        else:
                            modalidad = estudiante.modalidad.pk
                        
                        if str(estudiante.especialidad).isdigit():
                            especialidad =Especialidad.objects.get(pk=estudiante.especialidad)
                        else:
                            especialidad = estudiante.especialidad

                        if aprobacion:
                            inscripcion = Inscripcion(
                                curso=curso,
                                estudiante= estudiante,
                                especialidad = especialidad,
                                modalidad = ModalidadCursado.objects.get(pk = modalidad),
                                turno = Turno.objects.get(pk = turno),
                                estado = 'aprobado'
                                )
                        else:
                            inscripcion = Inscripcion(
                                curso=curso,
                                estudiante= estudiante,
                                especialidad = especialidad,
                                modalidad = ModalidadCursado.objects.get(pk = modalidad),
                                turno = Turno.objects.get(pk = turno),
                                estado = 'inscripto'
                                )
                        inscripcion.save()
                        mensajes.append('El documento {}, perteneciente a {}, {}, fue inscripto en {} exitosamente'.format(
                            docu,
                            persona.apellidos.upper(),
                            persona.nombres.title(),
                            curso
                            ))

                        email = create_email(
                            user_mail=estudiante.persona.correo,
                            subject='Confirmacion de la Inscripción',
                            template_name='correos/inscripcion/create.html',
                            context={
                                'nombre': estudiante.persona.nombres.title(),
                                'apellido': estudiante.persona.apellidos.upper(),
                                'carrera': inscripcion.especialidad.nombre.title(),
                                'turno': inscripcion.turno.nombre.title(),
                                'modalidad': inscripcion.modalidad.nombre.title(),
                                'seminario': inscripcion.curso.nombre.title(),
                                'id_estudiante': estudiante.pk,
                            },
                            request=request
                        )
                        thread = threading.Thread(target=email.send)
                        thread.start()
                    else:
                        inscripcion = Inscripcion.objects.filter(curso=curso).get(estudiante=estudiante)
                        if aprobacion:
                            inscripcion.estado = 'aprobado'
                        else:
                            inscripcion.estado='inscripto'
                        mensajes.append('El documento {}, perteneciente a {}, {}, ya se encontraba inscripto, su estado a sido actualizado a {}'.format(
                            docu,
                            persona.apellidos.upper(),
                            persona.nombres.title(),
                            inscripcion.estado.title(),
                        ))
                        inscripcion.save()
                            
                except Estudiante.DoesNotExist:
                    persona.delete()
                    mensajes.append('El documento {}, perteneciente a {}, {}, no cuenta con un estudiante, a sido borrado y debera completar la inscripcion nuevamente'.format(
                        docu,
                        persona.apellidos.upper(),
                        persona.nombres.title(),
                    ))
            except Persona.DoesNotExist:
                    mensajes.append('El documento {}, no existe'.format(
                        docu,
                    ))
        return JsonResponse({"success": True, "mensajes": mensajes})
    return JsonResponse({"success": False}, status=400)


class AdminHome(TemplateView, LoginRequiredMixin):
    template_name = "admin/home.html"
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        return context

